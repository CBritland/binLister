import re
import tkinter
from tkinter import StringVar, messagebox, END, SE
import ttkbootstrap as ttk
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import pymysql
import gspread
import os
from datetime import date

#This Program was built by Christopher Britland for any questions please contact Chris
#all sql information grabbed in this class
class Sql:
    # this function runs after user searches bin located in view_one in GUI class
    def initial_bin(self,bin1):
        sqlCon = pymysql.connect(host="10.1.10.41", port=3306, user="Admin", password="Seekonk123",
                                 database="inventory", charset='utf8mb4',
                                 cursorclass=pymysql.cursors.DictCursor)
        cur = sqlCon.cursor()
        cur.execute(f"SELECT `QR Num`, `Brand`, `Model`, `Sub Model`, `Condition Grade`, `Tested` FROM `chromebook inventory` WHERE `Bin` = '{bin1}' and `Listing Flag` = '0' and `Reset` = 'Yes'")
        qrFetch = cur.fetchall()
        sqlCon.close()
        GUI.build_tree(self, qrFetch)

    # this function runs after tree has been built and user selects a qr code to change to remove from bin
    def send_chris(self, qrNum):
        sqlCon = pymysql.connect(host="10.1.10.41", port=3306, user="Admin", password="Seekonk123",
                                 database="inventory", charset='utf8mb4',
                                 cursorclass=pymysql.cursors.DictCursor)
        cur = sqlCon.cursor()
        cur.execute(f"update `chromebook inventory` SET `Bin` = 'Chris' WHERE `QR Num` = '{qrNum}'")
        sqlCon.commit()
        sqlCon.close()

    # checks the bin to make sure the number in bin is correct this runs when user interacts with the enterButton
    def bin_check(self, binNum):
        print(binList)
        sqlCon = pymysql.connect(host="10.1.10.41", port=3306, user="Admin", password="Seekonk123",
                                database="inventory", charset='utf8mb4',
                                cursorclass=pymysql.cursors.DictCursor)
        cur = sqlCon.cursor()
        cur.execute(f"SELECT `Num` FROM `sortbins` WHERE `SortBin` = '{binNum}'")
        binFetch = cur.fetchall()
        for fetch in binFetch:
            for key,val in fetch.items():
                if key == 'Num':
                    numInBin = val
        if len(binList) != int(numInBin):
            cur2= sqlCon.cursor()
            cur2.execute(f"Update `sortbins` SET `Num` = '{len(binList)}' WHERE `SortBin` = '{binNum}'")
            sqlCon.commit()
            sqlCon.close()
        Listing.listing_lookup(self)

    # Looks to see if the selected model has a dd number correlated to a title
    def look_for_DD_in_models(self):
        global aDD
        global bDD
        global cDD
        global abcDD
        global dDD
        global eDD
        global mDD
        global demDD
        global fDD
        sqlCon = pymysql.connect(host="10.1.10.41", port=3306, user="Admin", password="Seekonk123",
                                 database="inventory", charset='utf8mb4',
                                 cursorclass=pymysql.cursors.DictCursor)
        cur = sqlCon.cursor()
        cur.execute(
            f"SELECT `A`, `B`, `C`, `ABC`, `D`, `E`, `M`, `DEM`, `F` FROM `chromebookmodels` WHERE `Brand` = '{Brand}' and `Model` = '{Model}' and `Sub Model` = '{subModel}'")
        ddFetch = cur.fetchall()
        print(73)
        print(ddFetch)
        for dd in ddFetch:
            for key, val in dd.items():
                if key == 'A':
                    aDD = val
                elif key == 'B':
                    bDD = val
                elif key == 'C':
                    cDD = val
                elif key == 'ABC':
                    abcDD = val
                elif key == 'D':
                    dDD = val
                elif key == 'E':
                    eDD = val
                elif key == 'M':
                    mDD = val
                elif key == 'DEM':
                    demDD = val
                elif key == 'F':
                    fDD = val
        Listing.check_info(self)

    # if grade is A B C D E F it comes here to pull the information from that one grade
    def sql_find_grades(self, gradeFind, numInGrade, DD, lots, singleInfo, description):
        sqlCon = pymysql.connect(host="10.1.10.41", port=3306, user="Admin", password="Seekonk123",
                                 database="inventory", charset='utf8mb4',
                                 cursorclass=pymysql.cursors.DictCursor)
        cur = sqlCon.cursor()
        cur.execute(f"SELECT * FROM `chromebook inventory` WHERE  `Brand` = '{Brand}' and `Model` = '{Model}' and `Sub Model` = '{subModel}' and `Bin` = '{binE.get()}' and `Listing Flag` = '0' and `Condition Grade` = '{gradeFind}'")
        gradesInfo = cur.fetchall()
        print(gradeFind)
        print(gradesInfo)
        Listing.grade_breakdown(self, gradesInfo, gradeFind, numInGrade, DD, lots, singleInfo, description)

    # if ABCLot or DEMLot it goes here to find the information before going to grade_breakdown in listing
    def multi_grade_find(self, gradeFind, numInGrade, DD, lots, singleInfo, description):
        if gradeFind == 'ABC':
            grade1 = 'A'
            grade2 = "B"
            grade3 = "C"
        elif gradeFind == 'DEM':
            grade1 = 'D'
            grade2 = 'E'
            grade3 = 'M'
        sqlCon = pymysql.connect(host="10.1.10.41", port=3306, user="Admin", password="Seekonk123",
                                 database="inventory", charset='utf8mb4',
                                 cursorclass=pymysql.cursors.DictCursor)
        cur = sqlCon.cursor()
        cur.execute(
            f"SELECT * FROM `chromebook inventory` WHERE `Brand` = '{Brand}' and `Model` = '{Model}' and `Sub Model` = '{subModel}' and `Bin` = '{binE.get()}' and `Listing Flag` = '0' and `Condition Grade` = '{grade1}' or `Brand` = '{Brand}' and `Model` = '{Model}' and `Sub Model` = '{subModel}' and `Bin` = '{binE.get()}' and `Listing Flag` = '0' and `Condition Grade` = '{grade2}' `Brand` = '{Brand}' and `Model` = '{Model}' and `Sub Model` = '{subModel}' and `Bin` = '{binE.get()}' and `Listing Flag` = '0' and `Condition Grade` = '{grade3}'")
        gradesInfo = cur.fetchall()
        print(gradeFind)
        print(gradesInfo)
        Listing.grade_breakdown(self, gradesInfo, gradeFind, numInGrade, DD, lots, singleInfo, description)
        
#-----------------------------------------------------------------------------------------------------------------------
# this class is for the listing portion of this program
class Listing:
    # connects to calculator to determine how many items belong in each
    def calculate_lookup(self):
        global singleA
        global singleB
        global singleC
        global singleD
        global singleE
        global singleM
        global singleF
        global lotACNum
        global lotDEMNum
        print(91)
        splitLotsList = []
        creds = gspread.service_account(filename=r"Z:\Chromebook Listing Tool\StephToolKey.json")
        sheet = creds.open("Lotting Calculator").worksheet('Sheet1')
        data = sheet.get_values()
        values_list = list(filter(None, sheet.col_values(3)))
        sheet.update_acell(f"C3", int(len(aList)))
        sheet.update_acell(f"C4", int(len(bList)))
        sheet.update_acell(f"C5", int(len(cList)))
        sheet.update_acell(f"C7", int(len(dList)))
        sheet.update_acell(f"C8", int(len(eList)))
        sheet.update_acell(f"C11", int(len(fList)))
        sheet.update_acell(f"C9", int(len(mList)))
        print(f'A{len(aList)}')
        print(f'B{len(bList)}')
        print(f'C{len(cList)}')
        print(f'D{len(dList)}')
        print(f'E{len(eList)}')
        print(f'{len(mList)}')
        print(f'{len(fList)}')
        singleA = sheet.cell(15, 3).value
        singleB = sheet.cell(16, 3).value
        singleC = sheet.cell(17, 3).value
        singleD = sheet.cell(20, 3).value
        singleE = sheet.cell(21, 3).value
        singleM = sheet.cell(22, 3).value
        singleF = sheet.cell(25, 3).value
        lotACNum = sheet.cell(18, 3).value
        lotDEMNum = sheet.cell(23, 3).value
        if int(singleA) != 0 and aDD != None:
            splitLotsList.append('singleA')
        if int(singleB) != 0 and bDD != None:
            splitLotsList.append('singleB')
        if int(singleC) != 0 and cDD != None:
            splitLotsList.append('singleC')
        if int(singleD) != 0 and dDD != None:
            splitLotsList.append('singleD')
        if int(singleE) != 0 and eDD != None:
            splitLotsList.append('singleE')
        if int(singleM) != 0 and mDD != None:
            splitLotsList.append('singleM')
        if int(singleF) != 0 and fDD != None:
            splitLotsList.append('singleF')
        if int(lotACNum) != 0 and abcDD != None:
            splitLotsList.append('lotACNum')
        if int(lotDEMNum) != 0 and demDD != None:
            splitLotsList.append('lotDEMNum')
        Listing.post_calculator(self, splitLotsList)

    # this function takes all the information sql_find_grades and breaks down the condition descrption
    def grade_breakdown(self, conditionGrab, gradeInformation, numInGrade, DD, lotList, single, description):
        problemList = []
        gradeListInfo = []
        permList = []
        info = 1
        newList = []
        firstList = True
        i = 0
        j=0
        listCheck = []
        eListChecker = []
        pathListing = 'Z:\Chromebook Listing Tool'
        creds = gspread.service_account(filename=r"Z:\ENTRY CSV\Inventory tool\Inventory.json")
        sheet = creds.open("New Inventory").worksheet('Inventory')
        os.chdir(pathListing)
        print(176)
        print(numInGrade)
        print(gradeInformation)
        print(os.listdir(pathListing))
        if f'bin{binE.get()}-{subModel}.xlsx' not in os.listdir(pathListing):
            print(177)
            wb = Workbook()
            ws = wb.active
            wb.create_sheet("Listings")
            ws1 = wb["Listings"]
            ws.cell(row=1, column=1).value = "DD"
            ws.cell(row=1, column=2).value = "QR"
            ws.cell(row=1, column=3).value = 'Grade'
            ws.cell(row=1, column=4).value = 'Description'
            wb.save(f'bin{binE.get()}-{subModel}.xlsx')
        else:
            wb = load_workbook(f'bin{binE.get()}-{subModel}.xlsx')
            ws = wb.active
            ws1 = wb["Listings"]
            ws.cell(row=ws.max_row + 1, column=1).value = ''
        print(conditionGrab)
        for condition in conditionGrab:
            for key, val in condition.items():
                if key == 'QR Num':
                    qr = val
                elif key == 'Cracked Screen':
                    if val == 'True':
                        problemList.append(key)
                elif key == 'Bad Backlight':
                    if val == 'True':
                        problemList.append(key)
                elif key == 'Has Water Damage':
                    if val == 'True':
                        problemList.append(key)
                elif key == 'Pressure Marks In Screen':
                    if val == 'True':
                        problemList.append(key)
                elif key == 'Lines In Display':
                    if val == 'True':
                        problemList.append(key)
                elif key == 'Screen Does Not Power On':
                    if val == 'True':
                        problemList.append(key)
                elif key == 'Screen Discoloration':
                    if val == 'True':
                        problemList.append(key)
                elif key == 'Screen Is Very Dim':
                    if val == 'True':
                        problemList.append('Dim')
                elif key == 'Screen Is Missing':
                    if val == 'True':
                        problemList.append(key)
                elif key == 'Screen Has Flicker':
                    if val == 'True':
                        problemList.append(key)
                elif key == 'Screen Is White':
                    if val == 'True':
                        problemList.append(key)
                elif key == 'Image Burnt Into Screen':
                    if val == 'True':
                        problemList.append(key)
                elif key == 'Screen Scratches':
                    if val == 'True':
                        problemList.append(key)
                elif key == 'Touch Pad Does Not Click':
                    if val == 'True':
                        problemList.append(key)
                elif key == 'Touch Pad No Movement':
                    if val == 'True':
                        problemList.append(key)
                elif key == 'Key Is Missing':
                    if val == 'True':
                        problemList.append(key)
                elif key == 'Key Does Not Work':
                    if val == 'True':
                        problemList.append(key)
                elif key == 'Scratch Grade':
                    if val == '2':
                        scratchProbD = 'Light Scratches'
                        problemList.append(scratchProbD)
                    if val == '3':
                        scratchProbD = 'Moderate Scratching'
                        problemList.append(scratchProbD)
                    if val == '4':
                        scratchProbD = 'Deep Scratching'
                        problemList.append(scratchProbD)
                elif key == 'Crack Grade':
                    if val == '3':
                        crackProbD = 'Small Crack Less Then 2 Inches'
                        problemList.append(crackProbD)
                    elif val == '4':
                        crackProbD = 'Crack More Then 2 Inches And/Or Multiple Cracks'
                        problemList.append(crackProbD)
                    elif val == '5':
                        crackProbD = 'Case Missing Plastic In One Or More Places'
                        problemList.append(crackProbD)
                elif key == 'Sticker Grade':
                    if val == '3':
                        stickerProbD = 'Device has engravings sticker, residue or markings less then 4 inches'
                        problemList.append(stickerProbD)
                    if val == '4':
                        stickerProbD = 'Device has engravings sticker, residue or markings more then 4 inches.'
                        problemList.append(stickerProbD)
                elif key == 'Dirty Grade':
                    if val == '4':
                        dirtyProb = 'Device Is Excessively Dirty'
                        problemList.append(dirtyProb)
                elif key == 'Port Grade':
                    if val == '3':
                        portProb = 'Missing Port Cover'
                        problemList.append(portProb)
                elif key == 'Feet Grade':
                    if val == '3':
                        feetProb = 'Missing Feet'
                        problemList.append(feetProb)
                elif key == 'ClientId':
                    print(i)
                    print(problemList)
                    dProbFix = re.sub('\[', '', str(problemList))
                    dProbFix1 = re.sub('\'', '', str(dProbFix))
                    dProbFix2 = re.sub(']', '', str(dProbFix1))
                    if i < int(numInGrade):
                        print(299)
                        ws.cell(row=ws.max_row + 1, column=1).value = DD
                        ws.cell(row=ws.max_row, column=2).value = qr
                        ws.cell(row=ws.max_row, column=3).value = gradeInformation
                        ws.cell(row=ws.max_row, column=4).value = str(dProbFix2)
                        permList.append(dProbFix2)
                        wb.save(f'bin{binE.get()}-{subModel}.xlsx')
                        problemList.clear()
                        i += 1
                        print(i)
                        print(numInGrade)
                        sqlCon1 = pymysql.connect(host="10.1.10.41", port=3306, user="Admin",
                                                  password="Seekonk123",
                                                  database="inventory", charset='utf8mb4',
                                                  cursorclass=pymysql.cursors.DictCursor)
                        cur1 = sqlCon1.cursor()
                        cur1.execute(
                            f"SELECT `Num` FROM `sortbins` WHERE `SortBin` = '{binE.get()}'")
                        aNumGrab = cur1.fetchall()
                        for num in aNumGrab:
                            for key, val in num.items():
                                if key == 'Num':
                                    binQty = val
                        newBinQty = int(binQty) - 1
                        if newBinQty <= 0:
                            cur2 = sqlCon1.cursor()
                            cur2.execute(
                                f"Update `sortbins` SET `Brand` = '', `Model` = '', `Sub Model` = '', `Num` = '', `Location` = '', `Device Type` = '', `Status` = ''  WHERE `SortBin` = '{binE.get()}'")
                        else:
                            cur2 = sqlCon1.cursor()
                            cur2.execute(
                                f"Update `sortbins` SET `Num` = '{newBinQty}' WHERE `SortBin` = '{binE.get()}'")
                        cur3 = sqlCon1.cursor()
                        cur3.execute(
                            f"Update `chromebook inventory` SET `Bin` = 'Listed', `Listing Flag` = 'Listed', `DD Number` = '{DD}' WHERE `QR Num` = '{qr}'")
                        sqlCon1.commit()
                        sqlCon1.close()
                        if i == int(numInGrade):
                            print(311)
                            print(lotList)
                            print(single)
                            lotList.remove(single)
                            Listing.post_calculator(self, lotList)


        for listInfoE in permList:
            for listInfoEAlt in permList:
                fixListInfoE = re.sub(']', '', str(listInfoE))
                fixListInfoEAlt = re.sub(']', '', str(listInfoEAlt))
                print(359)
                print(listCheck)
                print(fixListInfoE)
                print(fixListInfoEAlt)
                if fixListInfoE == fixListInfoEAlt:
                    print(363)
                    info += 1
                    newList.append(fixListInfoEAlt)
                    print(newList)
                    print(len(newList))
                else:
                    print(366)
            if firstList == True and fixListInfoE not in eListChecker:
                print(367)
                print(newList)
                print(len(newList))
                print(info)
                ws1.cell(row=ws1.max_row + 1, column=1).value = ''
                ws1.cell(row=ws1.max_row + 1, column=1).value = 'DD'
                ws1.cell(row=ws1.max_row, column=2).value = 'Grade'
                ws1.cell(row=ws1.max_row, column=3).value = 'QTY'
                ws1.cell(row=ws1.max_row, column=4).value = 'Description'
                ws1.cell(row=ws1.max_row, column=1).value = DD
                ws1.cell(row=ws1.max_row, column=2).value = gradeInformation
                ws1.cell(row=ws1.max_row, column=3).value = f'Total: {numInGrade}'
                ws1.cell(row=ws1.max_row + 1, column=3).value = len(newList)
                ws1.cell(row=ws1.max_row, column=4).value = str(fixListInfoE)
                wb.save(f'bin{binE.get()}-{subModel}.xlsx')
                eListChecker.append(fixListInfoE)
                ddFindE = sheet.find(DD)
                ddFixFindE = re.findall('\d+', str(ddFindE))[0]
                descriptionStringE = f'{description}\n{len(newList)} Units: {str(fixListInfoE)}'
                sheet.update_acell(f"G{int(ddFixFindE)}".format(), descriptionStringE)
                wb.close()
                eListChecker.append(fixListInfoE)
                firstList = False
                info = 1
                newList.clear()
            elif firstList == False and fixListInfoE not in eListChecker:
                print(396)
                descriptionStringE = f'{description}\n{len(newList)} Units: {str(fixListInfoE)}'
                sheet.update_acell(f"G{int(ddFixFindE)}".format(), descriptionStringE)
                ws1.cell(row=ws1.max_row + 1, column=3).value = len(newList)
                ws1.cell(row=ws1.max_row, column=4).value = str(fixListInfoE)
                eListChecker.append(fixListInfoE)
                wb.save(f'bin{binE.get()}-{subModel}.xlsx')
                wb.close()
                info = 1
                newList.clear()
            else:
                newList.clear()
        Listing.calculate_lookup(self)







    # this function determines whether titles have already been built for this item
    def check_info(self):
        if aDD == None and bDD == None and cDD == None and abcDD == None and dDD == None and eDD == None and mDD == None and demDD == None and fDD == None:
            print(117)
            GUI.clear(self)
        if aDD != None or bDD != None or cDD != None or abcDD != None or dDD != None or eDD != None or mDD != None or demDD != None:
            print(120)
            Listing.calculate_lookup(self)


    # this function grabs the information needed to see if titles have already been built for this item and is called in bin_check()
    def listing_lookup(self):
        global aDD
        global bDD
        global cDD
        global dDD
        global eDD
        global mDD
        global fDD
        global abcDD
        global demDD
        if len(noList) < 1:
            Sql.look_for_DD_in_models(self)
        else:
            errorResult = tkinter.messagebox.askyesno("ERROR!!", "There are chromebooks in bin that have not been tested. Would you like to proceed anyway?")
            if errorResult == True:
                Sql.look_for_DD_in_models(self)

    # after info is run through the calculator it comes here for further processing
    def post_calculator(self, lotsList):
        creds = gspread.service_account(filename=r"Z:\ENTRY CSV\Inventory tool\Inventory.json")
        sheet = creds.open("New Inventory").worksheet('Inventory')
        pathList = 'Z:\Chromebook Listing Tool'
        os.chdir(pathList)
        wb = load_workbook(r'Chromebook Listing Descriptions By Grade.xlsx', data_only=True)
        ws = wb['Sheet1']
        for lots in lotsList:
            print(380)
            print(lots)
            if lots == 'singleA':
                findSingleA = sheet.find(aDD)
                print(findSingleA)
                fixFindA = re.findall('\d+', str(findSingleA))[0]
                print(fixFindA)
                titleAFind = sheet.cell(str(fixFindA), 3).value
                print(titleAFind)
                abFind = sheet.cell(str(fixFindA), 28).value
                print(abFind)
                descriptionA = ws.cell(row=2, column=3).value
                wb.close()
                Sql.sql_find_grades(self, 'A', singleA, aDD, lotsList, 'singleA', descriptionA)
            if lots == 'singleB':
                findSingleB = sheet.find(bDD)
                print(findSingleB)
                fixFindB = re.findall('\d+', str(findSingleB))[0]
                print(fixFindB)
                titleBFind = sheet.cell(str(fixFindB), 3).value
                print(titleBFind)
                abFind = sheet.cell(str(fixFindB), 28).value
                print(abFind)
                descriptionB = ws.cell(row=3, column=3).value
                wb.close()
                Sql.sql_find_grades(self, 'B', singleB, bDD, lotsList, 'singleB', descriptionB)
            if lots == 'singleC':
                findSingleC = sheet.find(cDD)
                print(findSingleC)
                fixFindC = re.findall('\d+', str(findSingleC))[0]
                print(fixFindC)
                titleCFind = sheet.cell(str(fixFindC), 3).value
                print(titleCFind)
                abFind = sheet.cell(str(fixFindC), 28).value
                print(abFind)
                descriptionC = ws.cell(row=4, column=3).value
                wb.close()
                Sql.sql_find_grades(self, 'C', singleC, cDD, lotsList, 'singleC', descriptionC)
            if lots == 'singleD':
                findSingleD = sheet.find(dDD)
                print(findSingleD)
                fixFindD = re.findall('\d+', str(findSingleD))[0]
                print(fixFindD)
                titleDFind = sheet.cell(str(fixFindD), 3).value
                print(titleDFind)
                descriptionD = f'These ChromeBooks are used; They have been tested and reset using ChromeOS. They are all functional. They have been confirmed to be unmanaged. They are in poor physical condition, they may have one or more of the following problems: excessive scratches, deep gouges, large engraving, pieces of the case chipped or cracked away, larger cracks in the bodies, detached LCD, or might be excessively dirty . ' \
                                     'The units were used in a school and are in poor condition but are otherwise functional. The screens are not cracked.' \
                                     '\n\nLIST OF ITEMS IN LOT:\n'
                titleDFix1 = re.findall('(?<=Lot Of ).*', titleDFind)[0]
                titleDFix = re.findall('(?<= ).*', titleDFix1)[0]
                abFind = sheet.cell(str(fixFindD), 28).value
                print(abFind)
                print(titleDFix)
                if int(abFind) == 0:
                    titleD = f'Lot Of {singleD} {titleDFix}'
                    print(titleD)
                else:
                    afFind = sheet.cell(str(fixFindD), 32).value
                    newLotAmount = int(afFind)+int(singleD)
                    titleD = f'Lot of {newLotAmount} {titleDFix}'
                    print(titleD)
                Sql.sql_find_grades(self,'D', singleD, dDD, lotsList, 'singleD', descriptionD)
            if lots == 'singleE':
                findSingleE = sheet.find(eDD)
                print(findSingleE)
                fixFindE = re.findall('\d+', str(findSingleE))[0]
                print(fixFindE)
                titleEFind = sheet.cell(str(fixFindE), 3).value
                print(titleEFind)
                descriptionE = 'These ChromeBooks are used; They have been tested and reset using ChromeOS. They all have an issue (see below) and are being sold as-is. They have been confirmed to be unmanaged.The units were used in a school and the physical condition ranges from good to poor physical condition. The screens are not cracked.' \
                                     '\n\nLIST OF ITEMS IN LOT:\n'
                titleEFix1 = re.findall('(?<=Lot Of ).*', titleEFind)[0]
                titleEFix = re.findall('(?<= ).*', titleEFix1)[0]
                abFind = sheet.cell(str(fixFindE), 28).value
                print(abFind)
                print(titleEFix)
                if int(abFind) == 0:
                    titleD = f'Lot Of {singleE} {titleEFix}'
                    print(titleD)
                else:
                    afFind = sheet.cell(str(fixFindE), 32).value
                    newLotAmount = int(afFind) + int(singleE)
                    titleD = f'Lot of {newLotAmount} {titleEFix}'
                    print(titleD)
                Sql.sql_find_grades(self, 'E', singleE, eDD, lotsList, 'singleE', descriptionE)
            if lots == 'singleM':
                findSingleM = sheet.find(mDD)
                print(findSingleM)
                fixFindM = re.findall('\d+', str(findSingleM))[0]
                print(fixFindM)
                titleMFind = sheet.cell(str(fixFindM), 3).value
                print(titleMFind)
                titleMFix1 = re.findall('(?<=Lot Of ).*', titleMFind)[0]
                titleMFix = re.findall('(?<= ).*', titleMFix1)[0]
                abFind = sheet.cell(str(fixFindM), 28).value
                print(abFind)
                print(titleMFix)
                if int(abFind) == 0:
                    titleD = f'Lot Of {singleM} {titleMFix}'
                    print(titleD)
                else:
                    afFind = sheet.cell(str(fixFindM), 32).value
                    newLotAmount = int(afFind) + int(singleM)
                    titleD = f'Lot of {newLotAmount} {titleMFix}'
                    print(titleD)
                wb = load_workbook(r'Chromebook Listing Descriptions By Grade.xlsx', data_only=True)
                ws = wb['Sheet1']
                descriptionM = ws.cell(row=8, column=3).value
                wb.close()
                Sql.sql_find_grades(self, 'M', singleM, mDD, lotsList, 'singleM', descriptionM)
            if lots == 'singleF':
                findSingleF = sheet.find(fDD)
                print(findSingleF)
                fixFindF = re.findall('\d+', str(findSingleF))[0]
                print(fixFindF)
                titleFFind = sheet.cell(str(fixFindF), 3).value
                print(titleFFind)
                descriptionF = 'These ChromeBooks are used; They have been tested and reset using ChromeOS. They all have cracked screens and are being sold as-is. They have been confirmed to be unmanaged. The units were used in a school and the physical condition ranges from good to poor physical condition but are otherwise functional.' \
                                       '\n\nLIST OF ITEMS IN LOT:\n'
                titleFFix1 = re.findall('(?<=Lot Of ).*', titleFFind)[0]
                titleFFix = re.findall('(?<= ).*', titleFFix1)[0]
                abFind = sheet.cell(str(fixFindF), 28).value
                print(abFind)
                print(titleFFix)
                if int(abFind) == 0:
                    titleD = f'Lot Of {singleF} {titleFFix}'
                    print(titleD)
                else:
                    afFind = sheet.cell(str(fixFindF), 32).value
                    newLotAmount = int(afFind) + int(singleF)
                    titleD = f'Lot of {newLotAmount} {titleFFix}'
                    print(titleD)
                Sql.sql_find_grades(self, 'F', singleF, fDD, lotsList, 'singleF', descriptionF)
            if lots == 'lotACNum':
                findSingleABC = sheet.find(abcDD)
                print(findSingleABC)
                fixFindABC = re.findall('\d+', str(findSingleABC))[0]
                print(fixFindABC)
                titleAbcFind = sheet.cell(str(fixFindABC), 3).value
                print(titleAbcFind)
                descriptionABC = ws.cell(row=5, column=3).value
                wb.close()
                Sql.multi_grade_find(self, 'ABC', lotACNum, abcDD, lotsList, 'lotACNum', descriptionABC)

            if lots == 'lotDEMNum':
                findSingleDEM = sheet.find(demDD)
                print(findSingleDEM)
                fixFindDEM = re.findall('\d+', str(findSingleDEM))[0]
                print(fixFindDEM)
                titleDemFind = sheet.cell(str(fixFindDEM), 3).value
                titleDEMFix1 = re.findall('(?<=Lot Of ).*', titleDemFind)[0]
                titleDemFix = re.findall('(?<= ).*', titleDEMFix1)[0]
                print(titleDemFind)
                abFind = sheet.cell(str(fixFindDEM), 28).value
                wb = load_workbook(r'Chromebook Listing Descriptions By Grade.xlsx', data_only=True)
                ws = wb['Sheet1']
                if int(abFind) == 0:
                    titleD = f'Lot Of {lotDEMNum} {titleDemFix}'
                    print(titleD)
                else:
                    afFind = sheet.cell(str(fixFindDEM), 32).value
                    newLotAmount = int(afFind) + int(lotDEMNum)
                    titleD = f'Lot of {newLotAmount} {titleDemFix}'
                    print(titleD)
                descriptionDEM = ws.cell(row=9, column=3).value
                wb.close()
                Sql.multi_grade_find(self, 'DEM', lotDEMNum, demDD, lotsList, 'lotDEMNum', descriptionDEM)




# This Class is for building excel spreadsheet in event the user wants to find the chromebooks in question
#----------------------------------------------------------------------------------------------------------------------
class ExcelBuild:

    # building a spreadsheet based off grades
    def grade_build(self, gradeList):
        pathListing = 'Z:\Chromebook Listing Tool'
        os.chdir(pathListing)
        if f'{binE.get()}-{subModel}.xlsx' not in os.listdir(pathListing):
            wbs = Workbook()
            ws1 = wbs.active
            ws1.cell(row=1, column=1).value = "QR"
            ws1.cell(row=1, column=2).value = "Grade"
            ws1.cell(row=1, column=3).value = 'Description'
            wbs.save(f'{binE.get()}-{subModel}.xlsx')
        for grade in gradeList:
            print(grade)

    # asks user if they only want a sheet containing untested models
    def result_check(self, binInfo):
        resultFind = tkinter.messagebox.askyesno("Only No", "Would you like a sheet only containing untested chromebooks")
        ExcelBuild.build_sheet(self, binInfo, resultFind)

    # Builds out excel sheet triggered when user chooses to create sheet in
    def build_sheet(self, binInfo, result):
        today = date.today()
        todFix = str(today)
        global oldPath
        oldPath = os.getcwd()
        print(oldPath)
        newPath = r"Z:\binChecker\Reports"
        os.chdir(newPath)
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = 'Qr'
        ws.cell(row=1, column=2).value = 'Tested'
        ws.cell(row=1, column=3).value = "Grade"
        wb.save(f"bin{binE.get()}-{todFix}.xlsx")
        wb.close()
        print(165)
        print(result)
        for binQr in binInfo:
            for key,val in binQr.items():
                if key == 'QR Num':
                    qrExcel = val
                elif key == 'Condition Grade':
                    grade1 = val
                elif key == 'Tested':
                    testExcel = val
                    if result == True:
                        if testExcel == 'No':
                            wb1 = load_workbook(f"bin{binE.get()}-{todFix}.xlsx")
                            sheet = wb1.active
                            sheet.cell(row=sheet.max_row+1, column=1).value = qrExcel
                            sheet.cell(row=sheet.max_row, column=2).value = testExcel
                            sheet.cell(row=sheet.max_row, column=3).value = grade1
                            wb1.save(f"bin{binE.get()}-{todFix}.xlsx")
                            wb1.close()
                    else:
                        wb1 = load_workbook(f"bin{binE.get()}-{todFix}.xlsx")
                        sheet = wb1.active
                        sheet.cell(row=sheet.max_row + 1, column=1).value = qrExcel
                        sheet.cell(row=sheet.max_row, column=2).value = testExcel
                        sheet.cell(row=sheet.max_row, column=3).value = grade1
                        wb1.save(f"bin{binE.get()}-{todFix}.xlsx")
                        wb1.close()


#-----------------------------------------------------------------------------------------------------------------------
# this class builds out the GUI
class GUI:
    # does what it says it clears all the information out so the user can start over
    def clear(self):
        global Brand
        global Model
        global subModel
        global binList
        global aList
        global bList
        global cList
        global abcList
        global dList
        global eList
        global mList
        global demList
        global fList
        global aDD
        global bDD
        global cDD
        global dDD
        global eDD
        global mDD
        global fDD
        global abcDD
        global demDD
        global noList
        Brand = ''
        Model = ''
        noList = []
        subModel = ''
        binList.clear()
        aList.clear()
        bList.clear()
        cList.clear()
        dList.clear()
        eList.clear()
        mList.clear()
        abcList.clear()
        demList.clear()
        fList.clear()
        aDD = ''
        bDD = ''
        cDD = ""
        dDD = ''
        eDD = ''
        mDD = ''
        abcDD = ''
        demDD = ''
        fDD = ''
        binE.set('')
        treeFrame.destroy()
        underTree.destroy()

    # this is step 2 in the process this runs after initial_bin in sql
    def build_tree(self, qrInfo):
        global Brand
        global Model
        global subModel
        global treeFrame
        global underTree
        global noList
        noList = []
        treeFrame = ttk.Frame(root)
        treeFrame.grid(row=1, column=0, pady=10, sticky='w')
        underTree = ttk.Frame(root)
        underTree.grid(row=2, column=0, pady=10, sticky='w')
        columns = ('QR', 'Tested')
        tree = ttk.Treeview(treeFrame, bootstyle='danger', columns=columns, show='headings', height=40)
        tree.heading('QR', text='Qr')
        tree.heading('Tested', text='Tested')
        for qrGrab in qrInfo:
            for key,val in qrGrab.items():
                if key == 'QR Num':
                    qr = val
                    binList.append(qr)
                    print(qr)
                elif key == 'Brand':
                    Brand = val
                elif key == 'Model':
                    Model = val
                elif key == 'Sub Model':
                    subModel = val
                elif key == 'Condition Grade':
                    grade = val
                    if grade == 'A' or grade == 'a':
                        aList.append(qr)
                    elif grade == 'B' or grade == 'b':
                        bList.append(qr)
                    elif grade == 'C' or grade == 'c':
                        cList.append(qr)
                    elif grade == 'D' or grade == 'd':
                        dList.append(qr)
                    elif grade == 'E' or grade == 'e':
                        eList.append(qr)
                    elif grade == 'M' or grade == 'm':
                        mList.append(qr)
                    elif grade == 'F' or grade == 'f':
                        fList.append(qr)

                elif key == 'Tested':
                    test = val
                    if test == 'No':
                        noList.append(val)
                    print(test)
                    tree.insert('', END, values=(qr, test))
                    tree.grid(row=1,column=0)

        def treeview_sort_column(tv, col, reverse):
            l = [(tv.set(k, col), k) for k in tv.get_children('')]
            l.sort(reverse=reverse)

            # rearrange items in sorted positions
            for index, (val, k) in enumerate(l):
                tv.move(k, '', index)

            # reverse sort next time
            tv.heading(col, command=lambda: \
                treeview_sort_column(tv, col, not reverse))

        for col in columns:
            tree.heading(col, text=col, command=lambda: \
                treeview_sort_column(tree, col, False))

        #function runs when you choose a row in treeview
        def item_selected(event):
            global record
            for selected_item in tree.selection():
                item = tree.item(selected_item)
                record = item['values'][0]

        #confirms the removal of a chromebook from its bin
        def confirm_remove():
            try:
                result = tkinter.messagebox.askyesno("Testing Problem!!", f"Pass This {record} Onto Chris?")
                print(result)
                if result == True:
                    Sql.send_chris(self)
            except NameError:
                print(66)
                tkinter.messagebox.showerror("Choose Device", "You Must Select An Option From Spreadsheet View")

        tree.bind('<<TreeviewSelect>>', item_selected)
        removeFromBinLabel = ttk.Label(underTree, bootstyle='danger', text='Remove Item From Bin: ')
        removeFromBinLabel.configure(font=('Helvetica', 14))
        removeFromBinLabel.grid(row=0, column=0, sticky='w')
        removeFromBinButton = ttk.Button(underTree, bootstyle='danger-toolbutton', text='Enter', command=lambda: [confirm_remove()])
        removeFromBinButton.grid(row=0, column=1)
        createSpreadSheetLabel = ttk.Label(underTree, bootstyle='danger', text='Build Spreadsheet: ')
        createSpreadSheetLabel.configure(font=('Helvetica', 14))
        createSpreadSheetLabel.grid(row=1, column=0, sticky='w', pady=10)
        createSpreadSheetButton = ttk.Button(underTree, bootstyle='danger-toolbutton', text='Create', command=lambda: [ExcelBuild.result_check(self, qrInfo)])
        createSpreadSheetButton.grid(row=1, column=1)


    def __init__(self,root):
        # settting up basic gui layout
        self.root = None
        self.root = root
        self.root.title("Chromebook Bin Fixer")
        self.root.iconbitmap(r'Data.ico')
        ttk.Style("superhero")
        self.root.geometry("1350x780+0+0")
        global binList
        global aList
        global bList
        global cList
        global abcList
        global dList
        global eList
        global mList
        global demList
        global fList
        global binE
        binE = StringVar()
        binList = []
        aList = []
        bList = []
        cList = []
        abcList = []
        dList = []
        eList = []
        mList = []
        demList = []
        fList = []
        def view_one():
            global topFrame
            topFrame = ttk.Frame(root)
            topFrame.grid(row=0, column=0, sticky='w')
            binLabel = ttk.Label(topFrame, bootstyle='danger', text='Bin: ')
            binLabel.configure(font=('Helvetica', 14))
            binLabel.grid(row=0, column=0, sticky='w')
            binEntry = ttk.Entry(topFrame, bootstyle='danger', textvariable=binE)
            binEntry.grid(row=0, column=1)
            binEnter = ttk.Button(topFrame, bootstyle='danger-toolbutton', text='Enter', command=lambda: [Sql.initial_bin(self, binE.get())])
            binEnter.grid(row=0, column=2)
            binEntry.focus_set()
            binEntry.bind("<Return>", lambda event: Sql.initial_bin(self, binE.get()))
            enterButton = ttk.Button(root, bootstyle='danger-toolbutton', text='Enter', command=lambda: [Sql.bin_check(self, binE.get())])
            enterButton.place(rely=1.0, relx=1.0, x=0, y=0, anchor=SE)




        view_one()
if __name__ == '__main__':
    root = tkinter.Tk()
    application = GUI(root)
    root.mainloop()
